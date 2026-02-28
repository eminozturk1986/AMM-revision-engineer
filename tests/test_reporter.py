"""Tests for amm_engineer.reporter module."""

from pathlib import Path
import pytest
import openpyxl

from amm_engineer.differ import diff_revisions
from amm_engineer.parser import parse_text_amm
from amm_engineer.reporter import create_excel_report, create_markdown_report

SAMPLE_DIR = Path(__file__).parent.parent / "sample_data" / "amm_revisions"
REV15 = SAMPLE_DIR / "cessna172_amm_rev15_excerpt.txt"
REV21 = SAMPLE_DIR / "cessna172_amm_rev21_excerpt.txt"


@pytest.fixture
def changes():
    old_rev = parse_text_amm(REV15.read_text(encoding="utf-8"))
    new_rev = parse_text_amm(REV21.read_text(encoding="utf-8"))
    return diff_revisions(old_rev, new_rev)


@pytest.fixture
def amm_info():
    return {
        "aircraft": "Cessna 172",
        "old_rev": "15",
        "new_rev": "21",
        "date": "01 Jan 2026",
    }


def test_reporter_creates_excel(changes, amm_info, tmp_path):
    """Excel output must exist and have exactly 6 sheets."""
    out = str(tmp_path / "test_report.xlsx")
    result = create_excel_report(changes, out, amm_info)
    assert Path(result).exists(), f"Excel file not created at {result}"
    wb = openpyxl.load_workbook(result)
    assert len(wb.sheetnames) == 6, (
        f"Expected 6 sheets, got {len(wb.sheetnames)}: {wb.sheetnames}"
    )


def test_reporter_excel_sheet_names(changes, amm_info, tmp_path):
    """Excel sheets must have the correct names."""
    out = str(tmp_path / "test_names.xlsx")
    create_excel_report(changes, out, amm_info)
    wb = openpyxl.load_workbook(out)
    expected = {"SUMMARY", "ALL_CHANGES", "MRO_ACTION_ITEMS", "CAMO_AMP_IMPACT",
                "NEW_TASKS", "DELETED_TASKS"}
    assert set(wb.sheetnames) == expected, (
        f"Sheet names mismatch. Got: {wb.sheetnames}"
    )


def test_reporter_excel_has_data(changes, amm_info, tmp_path):
    """ALL_CHANGES sheet must have at least 1 data row."""
    out = str(tmp_path / "test_data.xlsx")
    create_excel_report(changes, out, amm_info)
    wb = openpyxl.load_workbook(out)
    ws = wb["ALL_CHANGES"]
    data_rows = ws.max_row - 1  # subtract header row
    assert data_rows >= 1, f"ALL_CHANGES sheet has no data rows (max_row={ws.max_row})"


def test_reporter_creates_markdown(changes, amm_info, tmp_path):
    """Markdown report must be created and contain key sections."""
    out = str(tmp_path / "test_report.md")
    result = create_markdown_report(changes, amm_info, out)
    assert Path(result).exists(), f"Markdown file not created at {result}"
    content = Path(result).read_text(encoding="utf-8")
    assert "AMM Revision Delta Report" in content
    assert "Priority" in content or "INTERVAL" in content

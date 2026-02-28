"""Generate Excel and Markdown reports from AMM revision analysis."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import AMMRevision, ChangeRecord

# ---------------------------------------------------------------------------
# Color palette (consistent with brand)
# ---------------------------------------------------------------------------
COLORS = {
    "header_bg": "1F4E79",
    "header_font": "FFFFFF",
    "high": "FF0000",
    "medium": "FF9900",
    "low": "00B050",
    "new_bg": "E2EFDA",
    "deleted_bg": "FFC7CE",
    "reduced_bg": "FFC7CE",
    "changed_bg": "FFEB9C",
    "zebra": "F2F2F2",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _header(ws, headers: list[str], row: int = 1) -> None:
    fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    font = Font(bold=True, color=COLORS["header_font"])
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(16, len(h) + 4)
    ws.row_dimensions[row].height = 28


def _row_fill(ws, row_idx: int, n_cols: int, hex_color: str) -> None:
    fill = PatternFill("solid", fgColor=hex_color)
    for col in range(1, n_cols + 1):
        ws.cell(row=row_idx, column=col).fill = fill


def _priority_color(priority: str) -> str:
    return {
        "HIGH": COLORS["high"],
        "MEDIUM": COLORS["medium"],
        "LOW": COLORS["low"],
    }.get(priority, "000000")


def _change_to_row(c: ChangeRecord) -> list[Any]:
    return [
        c.ata,
        c.task_ref,
        c.task_title,
        c.change_type,
        c.old_value or "",
        c.new_value or "",
        c.old_interval or "",
        c.new_interval or "",
        c.interval_delta or "",
        "; ".join(c.new_materials),
        "; ".join(c.new_tools),
        c.special_test or "",
        c.old_mh if c.old_mh is not None else "",
        c.new_mh if c.new_mh is not None else "",
        c.priority,
        c.action_required,
        c.effectivity,
    ]


_CHANGES_HEADERS = [
    "ATA", "Task_Ref", "Task_Title", "Change_Type",
    "Old_Value", "New_Value", "Old_Interval", "New_Interval",
    "Interval_Delta", "New_Materials", "New_Tools",
    "Special_Test", "Old_MH", "New_MH",
    "Priority", "Action_Required", "Effectivity",
]


def _build_changes_sheet(ws, changes: list[ChangeRecord]) -> None:
    _header(ws, _CHANGES_HEADERS)
    priority_col = _CHANGES_HEADERS.index("Priority") + 1

    for i, c in enumerate(changes, 2):
        row_data = _change_to_row(c)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Row background
        ct = c.change_type
        if ct == "INTERVAL_REDUCED":
            _row_fill(ws, i, len(_CHANGES_HEADERS), COLORS["reduced_bg"])
        elif ct == "NEW_TASK":
            _row_fill(ws, i, len(_CHANGES_HEADERS), COLORS["new_bg"])
        elif ct == "DELETED_TASK":
            _row_fill(ws, i, len(_CHANGES_HEADERS), COLORS["deleted_bg"])
        elif i % 2 == 0:
            _row_fill(ws, i, len(_CHANGES_HEADERS), COLORS["zebra"])

        # Priority cell colour
        ws.cell(row=i, column=priority_col).font = Font(
            color=_priority_color(c.priority), bold=True
        )

    if changes:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _build_summary_sheet(ws, changes: list[ChangeRecord], amm_info: dict) -> None:
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12

    ws["A1"] = f"AMM Revision Delta Report — {amm_info.get('aircraft', '')}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = f"Revision {amm_info.get('old_rev', '?')} → Revision {amm_info.get('new_rev', '?')}"
    ws["A2"].font = Font(size=11)
    ws["A3"] = f"Analysis Date: {amm_info.get('date', datetime.now().strftime('%d %b %Y'))}"
    ws["A3"].font = Font(italic=True, color="595959")

    ws["A5"] = "Change Type"
    ws["B5"] = "Count"
    ws["A5"].font = Font(bold=True)
    ws["B5"].font = Font(bold=True)

    stats: dict[str, int] = {}
    priority_counts: dict[str, int] = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
    for c in changes:
        stats[c.change_type] = stats.get(c.change_type, 0) + 1
        priority_counts[c.priority] = priority_counts.get(c.priority, 0) + 1

    row = 6
    for ct, count in sorted(stats.items()):
        cell_a = ws.cell(row=row, column=1, value=ct)
        cell_b = ws.cell(row=row, column=2, value=count)
        if ct == "INTERVAL_REDUCED":
            cell_a.font = Font(color=COLORS["high"], bold=True)
            cell_b.font = Font(color=COLORS["high"], bold=True)
        row += 1

    ws.cell(row=row, column=1, value="TOTAL CHANGES").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(changes)).font = Font(bold=True)
    row += 2

    ws.cell(row=row, column=1, value="Priority Breakdown").font = Font(bold=True)
    row += 1
    for prio, count in priority_counts.items():
        ws.cell(row=row, column=1, value=prio).font = Font(
            color=_priority_color(prio), bold=True
        )
        ws.cell(row=row, column=2, value=count)
        row += 1


def _build_camo_sheet(ws, changes: list[ChangeRecord]) -> None:
    relevant_types = {
        "INTERVAL_REDUCED", "INTERVAL_EXTENDED", "INTERVAL_CHANGE",
        "NEW_TASK", "DELETED_TASK",
    }
    camo = [c for c in changes if c.change_type in relevant_types]

    headers = [
        "ATA", "Task_Ref", "Task_Title", "AMP_Task_Match",
        "Old_Interval", "New_Interval", "Interval_Delta",
        "Change_Type", "Authority_Notification", "AMP_Amendment_Required",
        "Action_Required", "Priority",
    ]
    _header(ws, headers)

    for i, c in enumerate(camo, 2):
        auth_notif = "YES" if c.change_type == "INTERVAL_REDUCED" else "EVALUATE"
        amp_amend = "YES" if c.change_type in (
            "INTERVAL_REDUCED", "INTERVAL_EXTENDED", "NEW_TASK", "DELETED_TASK"
        ) else "NO"

        row_data = [
            c.ata, c.task_ref, c.task_title,
            c.amp_task_match or "Not matched",
            c.old_interval or "", c.new_interval or "",
            c.interval_delta or "", c.change_type,
            auth_notif, amp_amend,
            c.action_required, c.priority,
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        if c.change_type == "INTERVAL_REDUCED":
            _row_fill(ws, i, len(headers), COLORS["reduced_bg"])
        elif c.change_type == "NEW_TASK":
            _row_fill(ws, i, len(headers), COLORS["new_bg"])

        priority_col = headers.index("Priority") + 1
        ws.cell(row=i, column=priority_col).font = Font(
            color=_priority_color(c.priority), bold=True
        )

    if camo:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_excel_report(
    changes: list[ChangeRecord],
    output_path: str,
    amm_info: dict,
) -> str:
    """
    Create full Excel AMM revision delta report with 6 sheets.

    Args:
        changes: List of ChangeRecord objects from diff_revisions().
        output_path: Destination .xlsx file path.
        amm_info: {'aircraft', 'old_rev', 'new_rev', 'date'}

    Returns:
        Absolute path of saved file.
    """
    wb = openpyxl.Workbook()

    # Sheet 1: SUMMARY
    ws_sum = wb.active
    ws_sum.title = "SUMMARY"
    _build_summary_sheet(ws_sum, changes, amm_info)

    # Sheet 2: ALL_CHANGES
    ws_all = wb.create_sheet("ALL_CHANGES")
    _build_changes_sheet(ws_all, changes)

    # Sheet 3: MRO_ACTION_ITEMS
    mro_types = {
        "NEW_TOOL", "TOOL_CHANGED", "NEW_MATERIAL", "MATERIAL_CHANGED",
        "NEW_SPECIAL_TEST", "MH_CHANGE", "NEW_TASK", "DELETED_TASK",
    }
    ws_mro = wb.create_sheet("MRO_ACTION_ITEMS")
    _build_changes_sheet(ws_mro, [c for c in changes if c.change_type in mro_types])

    # Sheet 4: CAMO_AMP_IMPACT
    ws_camo = wb.create_sheet("CAMO_AMP_IMPACT")
    _build_camo_sheet(ws_camo, changes)

    # Sheet 5: NEW_TASKS
    ws_new = wb.create_sheet("NEW_TASKS")
    _build_changes_sheet(ws_new, [c for c in changes if c.change_type == "NEW_TASK"])

    # Sheet 6: DELETED_TASKS
    ws_del = wb.create_sheet("DELETED_TASKS")
    _build_changes_sheet(ws_del, [c for c in changes if c.change_type == "DELETED_TASK"])

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out.resolve())


def create_markdown_report(
    changes: list[ChangeRecord],
    amm_info: dict,
    output_path: str,
) -> str:
    """
    Write a Markdown summary report alongside the Excel output.

    Returns path of created file.
    """
    high = [c for c in changes if c.priority == "HIGH"]
    medium = [c for c in changes if c.priority == "MEDIUM"]
    low = [c for c in changes if c.priority == "LOW"]

    interval_reduced = [c for c in changes if c.change_type == "INTERVAL_REDUCED"]
    new_tasks = [c for c in changes if c.change_type == "NEW_TASK"]
    deleted = [c for c in changes if c.change_type == "DELETED_TASK"]
    new_tests = [c for c in changes if c.change_type == "NEW_SPECIAL_TEST"]
    new_tools = [c for c in changes if c.change_type == "NEW_TOOL"]
    new_mats = [c for c in changes if c.change_type in ("NEW_MATERIAL", "MATERIAL_CHANGED")]
    mh = [c for c in changes if c.change_type == "MH_CHANGE"]

    lines = [
        f"# AMM Revision Delta Report",
        f"",
        f"**Aircraft:** {amm_info.get('aircraft', 'N/A')}  ",
        f"**AMM:** Rev {amm_info.get('old_rev', '?')} → Rev {amm_info.get('new_rev', '?')}  ",
        f"**Analysis Date:** {amm_info.get('date', datetime.now().strftime('%d %b %Y'))}",
        f"",
        f"---",
        f"",
        f"## Summary",
        f"",
        f"| Priority | Count |",
        f"|----------|-------|",
        f"| HIGH     | {len(high)} |",
        f"| MEDIUM   | {len(medium)} |",
        f"| LOW      | {len(low)} |",
        f"| **TOTAL**| **{len(changes)}** |",
        f"",
        f"---",
        f"",
        f"## Priority Actions",
        f"",
    ]

    if interval_reduced:
        lines.append(f"### Interval Reductions (AMP Amendment Required)")
        for c in interval_reduced:
            lines.append(
                f"- **{c.ata} — {c.task_title}**: "
                f"{c.old_interval} → {c.new_interval} "
                f"({c.interval_delta})  "
            )
            if c.amp_task_match:
                lines.append(f"  AMP Task: `{c.amp_task_match}`")
        lines.append("")

    if new_tests:
        lines.append("### New Special Tests (Equipment / Personnel Required)")
        for c in new_tests:
            lines.append(f"- **{c.ata} — {c.task_title}**: {c.special_test}")
        lines.append("")

    if new_tasks:
        lines.append("### New Tasks (Engineering Evaluation Required)")
        for c in new_tasks:
            ivl = c.new_interval or "N/A"
            lines.append(f"- **{c.ata} — {c.task_title}** (Interval: {ivl})")
        lines.append("")

    if deleted:
        lines.append("### Deleted Tasks")
        for c in deleted:
            lines.append(f"- **{c.ata} — {c.task_title}**")
        lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## MRO Impact")
    lines.append("")

    if new_tools:
        lines.append("### New Tools Required")
        for c in new_tools:
            lines.append(f"- {c.ata}: {'; '.join(c.new_tools)}")
        lines.append("")

    if new_mats:
        lines.append("### Material Changes")
        for c in new_mats:
            if c.change_type == "MATERIAL_CHANGED":
                lines.append(f"- {c.ata} — {c.task_title}: `{c.old_value}` → `{c.new_value}`")
            else:
                lines.append(f"- {c.ata} — {c.task_title}: {', '.join(c.new_materials)}")
        lines.append("")

    if mh:
        lines.append("### Man-Hour Changes")
        for c in mh:
            lines.append(
                f"- {c.ata} — {c.task_title}: {c.old_mh} MH → {c.new_mh} MH"
            )
        lines.append("")

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")
    return str(out.resolve())
